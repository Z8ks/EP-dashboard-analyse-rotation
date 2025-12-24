import os
import glob
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOSSIER_DATA = r"F:\Data"
DOSSIER_SORTIE = r"F:\02_Analyse_Rotation\Dashboard"
ENSEIGNE = "ELECTROPLANET"


def safe_read_excel(file_path, sheet_name=0, **kwargs):
    """Lecture Excel robuste"""
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)
    except Exception as e:
        print(f"Erreur lecture: {e}")
        return pd.DataFrame()


def appliquer_style_header(ws, row, color_hex, bold=True):
    """Style pour les headers"""
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    font = Font(color="FFFFFF", bold=bold, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def appliquer_couleur_status(ws, col_letter, start_row, end_row):
    """Applique des couleurs basées sur le STATUS"""
    couleurs_status = {
        "URGENT": "FF0000",
        "COMMANDE": "FFA500",
        "BLOCKBUSTER": "00B050",
        "MORT": "FF6B35",
        "STABLE": "D3D3D3"
    }
    
    for row in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{row}"]
        text = str(cell.value)
        for keyword, color in couleurs_status.items():
            if keyword in text:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                break


def generer_recommandations(dashboard, week_num):
    """Génère des recommandations hebdomadaires basées sur les données"""
    
    nb_urgents = len(dashboard[dashboard['COUVERTURE'] < 14])
    nb_a_commander = len(dashboard[(dashboard['COUVERTURE'] >= 14) & (dashboard['COUVERTURE'] < 28)])
    nb_stock_mort = len(dashboard[(dashboard['VENTES_HEBDO'] == 0) & (dashboard['STOCK_EP'] > 10)])
    nb_blockbusters = len(dashboard[dashboard['ROTATION'] > 0.5])
    
    ca_total = dashboard['CA_HEBDO'].sum()
    stock_total_ep = dashboard['STOCK_EP'].sum()
    ventes_totales = dashboard['VENTES_HEBDO'].sum()
    
    top_urgents = dashboard[dashboard['COUVERTURE'] < 14].nlargest(3, 'CA_HEBDO')[['LIBELLE', 'STOCK_EP', 'VENTES_HEBDO', 'COUVERTURE']]
    
    stock_mort = dashboard[(dashboard['VENTES_HEBDO'] == 0) & (dashboard['STOCK_EP'] > 10)].copy()
    stock_mort['VALEUR_IMMOBILISEE'] = stock_mort['STOCK_EP'] * stock_mort['P.ACHAT']
    top_stock_mort = stock_mort.nlargest(3, 'VALEUR_IMMOBILISEE')[['LIBELLE', 'STOCK_EP', 'VALEUR_IMMOBILISEE']]
    
    top_blockbusters = dashboard[dashboard['ROTATION'] > 0.5].nlargest(3, 'CA_HEBDO')[['LIBELLE', 'CA_HEBDO', 'ROTATION']]
    
    recommandations = []
    
    recommandations.append(["VUE D'ENSEMBLE", ""])
    recommandations.append(["Semaine analysee", f"W{week_num}"])
    recommandations.append(["CA hebdomadaire", f"{ca_total:,.0f} DH"])
    recommandations.append(["Ventes totales", f"{ventes_totales:,.0f} unites"])
    recommandations.append(["Stock total EP", f"{stock_total_ep:,.0f} unites"])
    recommandations.append(["", ""])
    
    recommandations.append(["ACTIONS PRIORITAIRES", ""])
    
    if nb_urgents > 0:
        recommandations.append(["[URGENT]", f"{nb_urgents} articles en risque de rupture"])
        recommandations.append(["Action requise", "Commander IMMEDIATEMENT ces articles (< 14 jours de stock)"])
        if not top_urgents.empty:
            recommandations.append(["Top 3 urgents", ""])
            for idx, row in top_urgents.iterrows():
                lib = str(row['LIBELLE'])[:40]
                recommandations.append(["", f"- {lib} | Stock: {row['STOCK_EP']:.0f} | Ventes: {row['VENTES_HEBDO']:.0f} | Couvre: {row['COUVERTURE']:.1f}j"])
    else:
        recommandations.append(["[URGENT]", "Aucun article en rupture imminente"])
    
    recommandations.append(["", ""])
    
    if nb_a_commander > 0:
        recommandations.append(["[A COMMANDER]", f"{nb_a_commander} articles a prevoir"])
        recommandations.append(["Action requise", "Planifier commande dans les 7 prochains jours (14-28 jours de stock)"])
    else:
        recommandations.append(["[A COMMANDER]", "Stock bien gere"])
    
    recommandations.append(["", ""])
    
    recommandations.append(["OPPORTUNITES", ""])
    
    if nb_blockbusters > 0:
        recommandations.append(["[BLOCKBUSTERS]", f"{nb_blockbusters} produits stars (rotation > 50%)"])
        recommandations.append(["Action requise", "Augmenter le stock de ces articles a forte demande"])
        if not top_blockbusters.empty:
            recommandations.append(["Top 3 performers", ""])
            for idx, row in top_blockbusters.iterrows():
                lib = str(row['LIBELLE'])[:40]
                recommandations.append(["", f"- {lib} | CA: {row['CA_HEBDO']:,.0f} DH | Rotation: {row['ROTATION']:.2f}"])
    else:
        recommandations.append(["[BLOCKBUSTERS]", "Aucun produit a forte rotation"])
    
    recommandations.append(["", ""])
    
    recommandations.append(["PROBLEMES A RESOUDRE", ""])
    
    if nb_stock_mort > 0:
        valeur_immobilisee = stock_mort['VALEUR_IMMOBILISEE'].sum()
        recommandations.append(["[STOCK MORT]", f"{nb_stock_mort} articles sans vente"])
        recommandations.append(["Valeur immobilisee", f"{valeur_immobilisee:,.0f} DH"])
        recommandations.append(["Action requise", "Lancer promotions / destockage / retour fournisseur"])
        if not top_stock_mort.empty:
            recommandations.append(["Top 3 a destocquer", ""])
            for idx, row in top_stock_mort.iterrows():
                lib = str(row['LIBELLE'])[:40]
                recommandations.append(["", f"- {lib} | Stock: {row['STOCK_EP']:.0f} | Valeur: {row['VALEUR_IMMOBILISEE']:,.0f} DH"])
    else:
        recommandations.append(["[STOCK MORT]", "Aucun article problematique"])
    
    recommandations.append(["", ""])
    
    recommandations.append(["PLAN D'ACTION CETTE SEMAINE", ""])
    
    action_num = 1
    if nb_urgents > 0:
        recommandations.append([f"Action {action_num}", f"Commander {nb_urgents} articles urgents avant rupture"])
        action_num += 1
    
    if nb_a_commander > 0:
        recommandations.append([f"Action {action_num}", f"Preparer commande de {nb_a_commander} articles (prevoir delai livraison)"])
        action_num += 1
    
    if nb_stock_mort > 0:
        recommandations.append([f"Action {action_num}", f"Lancer promotion sur {min(nb_stock_mort, 10)} articles sans rotation"])
        action_num += 1
    
    if nb_blockbusters > 0:
        recommandations.append([f"Action {action_num}", f"Augmenter stock des {nb_blockbusters} blockbusters pour maximiser CA"])
        action_num += 1
    
    if action_num == 1:
        recommandations.append(["Action 1", "Gestion normale - Surveillance continue"])
    
    return recommandations


# FICHIERS
print("="*80)
print("LANCEMENT GENERATION DASHBOARD RETAILER")
print("="*80)

stock_files = glob.glob(os.path.join(DOSSIER_DATA, "ExcelStock-*.xlsx"))
vente_files = glob.glob(os.path.join(DOSSIER_DATA, "ExcelVenteHebdo-*.xlsx"))
recap_files = [f for f in glob.glob(os.path.join(DOSSIER_DATA, "*RECAP*.xlsx")) if not os.path.basename(f).startswith('~$')]
burintel_files = glob.glob(os.path.join(DOSSIER_DATA, "*LABBURINTEL*.xlsx"))

stock_file = max(stock_files, key=os.path.getctime) if stock_files else None
vente_file = max(vente_files, key=os.path.getctime) if vente_files else None
recap_file = max(recap_files, key=os.path.getctime) if recap_files else None
burintel_file = max(burintel_files, key=os.path.getctime) if burintel_files else None

print(f"\nFichiers detectes:")
print(f"   Stock: {os.path.basename(stock_file) if stock_file else 'INTROUVABLE'}")
print(f"   Ventes: {os.path.basename(vente_file) if vente_file else 'INTROUVABLE'}")
print(f"   RECAP: {os.path.basename(recap_file) if recap_file else 'INTROUVABLE'}")
print(f"   Burintel: {os.path.basename(burintel_file) if burintel_file else 'INTROUVABLE'}")

# CHARGEMENT
df_stock = safe_read_excel(stock_file, "Stock") if stock_file else pd.DataFrame()
df_ventes = safe_read_excel(vente_file, "Ventes hebdomadaires") if vente_file else pd.DataFrame()
df_burintel = safe_read_excel(burintel_file, 0) if burintel_file else pd.DataFrame()
df_recap = safe_read_excel(recap_file, 0) if recap_file else pd.DataFrame()

# FALLBACK
if df_recap.empty and not df_stock.empty:
    print("RECAP vide -> Utilisation du Stock EP filtre")
    df_stock.columns = [str(c).strip() for c in df_stock.columns]
    
    if "Enseigne" in df_stock.columns:
        df_recap = df_stock[df_stock["Enseigne"] == ENSEIGNE].copy()
    else:
        df_recap = df_stock.copy()
    
    df_recap = df_recap.rename(columns={
        "Libellé article": "Libelle EP",
        "Quantité": "Stock EP"
    })
    print(f"{len(df_recap)} articles Stock EP utilises comme RECAP")

if df_recap.empty:
    print("AUCUNE DONNEE DISPONIBLE -> EXIT")
    input("Appuyez sur Entree...")
    exit()

# NORMALISATION
df_recap.columns = [str(c).strip() for c in df_recap.columns]
df_ventes.columns = [str(c).strip() for c in df_ventes.columns]
if not df_burintel.empty:
    df_burintel.columns = [str(c).strip() for c in df_burintel.columns]

for df in [df_recap, df_ventes, df_stock, df_burintel]:
    if isinstance(df, pd.DataFrame) and not df.empty and "EAN" in df.columns:
        df["EAN"] = df["EAN"].astype(str).str.strip()

print(f"\n{len(df_recap)} articles charges dans RECAP")
print(f"{len(df_ventes)} lignes de ventes chargees")
if not df_burintel.empty:
    print(f"{len(df_burintel)} articles Burintel charges")

# STOCK BURINTEL
burintel_stock = pd.DataFrame()
if not df_burintel.empty and "Stock Burintel" in df_burintel.columns:
    burintel_stock = df_burintel[["N°", "Description", "Stock Burintel"]].copy()
    burintel_stock.columns = ["N°", "Description", "STOCK_BURINTEL"]
    burintel_stock["STOCK_BURINTEL"] = pd.to_numeric(burintel_stock["STOCK_BURINTEL"], errors='coerce').fillna(0)
    print(f"Stock Burintel: {burintel_stock['STOCK_BURINTEL'].sum():,.0f} unites")

# VENTES EP
if "Libellé Enseigne" in df_ventes.columns:
    df_ventes_ep = df_ventes[df_ventes["Libellé Enseigne"] == ENSEIGNE].copy()
elif "Enseigne" in df_ventes.columns:
    df_ventes_ep = df_ventes[df_ventes["Enseigne"] == ENSEIGNE].copy()
else:
    df_ventes_ep = df_ventes.copy()

# SEMAINE
date_semaine = pd.Timestamp.now()
df_ventes_sem = pd.DataFrame()

if not df_ventes_ep.empty and "Début semaine" in df_ventes_ep.columns:
    df_ventes_ep["Début semaine"] = pd.to_datetime(df_ventes_ep["Début semaine"], dayfirst=True, errors='coerce')
    date_semaine = df_ventes_ep["Début semaine"].max()
    df_ventes_sem = df_ventes_ep[df_ventes_ep["Début semaine"] == date_semaine].copy()

week_num = date_semaine.isocalendar()[1]
print(f"\nSemaine W{week_num}: {date_semaine.strftime('%d/%m/%Y')}")
print(f"{len(df_ventes_sem)} lignes de ventes pour cette semaine")

# VENTES PAR EAN
if not df_ventes_sem.empty and "Quantité" in df_ventes_sem.columns:
    ventes_ean = df_ventes_sem.groupby("EAN")["Quantité"].sum().reset_index(name="VENTES_HEBDO")
    print(f"{len(ventes_ean)} EAN avec ventes")
else:
    ventes_ean = pd.DataFrame({"EAN": df_recap["EAN"].unique(), "VENTES_HEBDO": 0})
    print("Pas de ventes hebdo - Initialisation a 0")

# DASHBOARD
dashboard = df_recap.merge(ventes_ean, on="EAN", how="left").fillna({"VENTES_HEBDO": 0})

# COLONNES
dashboard["MARQUE"] = dashboard.get("MARQUE", dashboard.get("Code Burintel", "NC")).fillna("NC")
dashboard["LIBELLE"] = dashboard.get("Libelle EP", dashboard.get("Libellé article", "Article")).fillna("Article")
dashboard["STOCK_EP"] = pd.to_numeric(dashboard.get("Stock EP", dashboard.get("Quantité", 0)), errors='coerce').fillna(0)
dashboard["P.VENTE"] = pd.to_numeric(dashboard.get("P.Vente", 0), errors='coerce').fillna(0)
dashboard["P.ACHAT"] = pd.to_numeric(dashboard.get("P.Achat", 0), errors='coerce').fillna(0)
dashboard["BURINTEL_DEPOT"] = pd.to_numeric(dashboard.get("BURINTEL DEPOT", 0), errors='coerce').fillna(0)

# KPI
dashboard["CA_HEBDO"] = dashboard["VENTES_HEBDO"] * dashboard["P.VENTE"]
dashboard["STOCK_TOTAL"] = dashboard["STOCK_EP"] + dashboard["BURINTEL_DEPOT"]
dashboard["ROTATION"] = np.divide(dashboard["VENTES_HEBDO"], dashboard["STOCK_EP"], 
                                   where=dashboard["STOCK_EP"]>0, 
                                   out=np.zeros_like(dashboard["VENTES_HEBDO"], dtype=float)).round(2)
dashboard["COUVERTURE"] = np.where(dashboard["VENTES_HEBDO"] > 0, 
                                   dashboard["STOCK_EP"] / dashboard["VENTES_HEBDO"] * 7, 999).round(1)

# STATUS (SANS EMOJIS - on les ajoutera après)
conditions = [
    (dashboard["VENTES_HEBDO"] == 0) & (dashboard["STOCK_EP"] > 10),
    dashboard["COUVERTURE"] < 14,
    (dashboard["COUVERTURE"] >= 14) & (dashboard["COUVERTURE"] < 28),
    dashboard["ROTATION"] > 0.5
]
choices = ["STOCK MORT", "URGENT", "A COMMANDE", "BLOCKBUSTER"]
dashboard["STATUS"] = np.select(conditions, choices, default="STABLE")

print(f"\nINDICATEURS:")
print(f"   Urgents: {len(dashboard[dashboard['COUVERTURE'] < 14])}")
print(f"   A commander: {len(dashboard[(dashboard['COUVERTURE'] >= 14) & (dashboard['COUVERTURE'] < 28)])}")
print(f"   Blockbusters: {len(dashboard[dashboard['ROTATION'] > 0.5])}")
print(f"   Stock mort: {len(dashboard[(dashboard['VENTES_HEBDO'] == 0) & (dashboard['STOCK_EP'] > 10)])}")

# TOPS
medailles = ["#1", "#2", "#3"] + [f"#{i}" for i in range(4,11)]
top_ca = dashboard.nlargest(10, "CA_HEBDO")[["MARQUE", "LIBELLE", "CA_HEBDO", "VENTES_HEBDO", "STATUS"]]
if not top_ca.empty:
    top_ca = top_ca.copy()
    top_ca.insert(0, "Rang", medailles[:len(top_ca)])
    top_ca.columns = ["Rang", "MARQUE", "Article", "CA", "Ventes", "Status"]

# KPI (SANS EMOJIS)
kpi_data = {
    f"CA W{week_num}": f"{dashboard['CA_HEBDO'].sum():,.0f} DH",
    "Stock EP": f"{dashboard['STOCK_EP'].sum():,.0f} unites",
    "Burintel Depot": f"{dashboard['BURINTEL_DEPOT'].sum():,.0f} unites",
    "Ventes Hebdo": f"{dashboard['VENTES_HEBDO'].sum():,.0f} unites",
    "Articles Urgents": f"{len(dashboard[dashboard['COUVERTURE'] < 14])} articles",
    "Blockbusters": f"{len(dashboard[dashboard['ROTATION'] > 0.5])} articles"
}

# 📚 GUIDE DE LECTURE (CORRECTION DES FORMULES)
guide_lecture = [
    ["GUIDE DE LECTURE - COMPRENDRE LE DASHBOARD", ""],
    ["", ""],
    ["FORMULES UTILISEES", ""],
    ["ROTATION", "'= Ventes Hebdo / Stock EP"],  # ✅ Apostrophe ajoutée
    ["", "Mesure: Combien de fois le stock tourne par semaine"],
    ["", "Exemple: Rotation 0.50 = 50% du stock vendu chaque semaine"],
    ["", ""],
    ["COUVERTURE", "'= (Stock EP / Ventes Hebdo) x 7 jours"],  # ✅ Apostrophe ajoutée
    ["", "Mesure: Nombre de jours avant rupture au rythme actuel"],
    ["", "Exemple: Couverture 14 jours = 2 semaines de stock"],
    ["", ""],
    ["CA HEBDO", "'= Ventes Hebdo x Prix de Vente"],  # ✅ Apostrophe ajoutée
    ["", "Mesure: Chiffre d'affaires genere cette semaine"],
    ["", ""],
    ["SIGNIFICATION DES STATUS", ""],
    ["[URGENT]", "Couverture < 14 jours -> Risque de rupture immediate"],
    ["Action", "Commander IMMEDIATEMENT avant rupture de stock"],
    ["", ""],
    ["[A COMMANDE]", "Couverture entre 14 et 28 jours -> Stock normal"],
    ["Action", "Planifier commande dans les 7 prochains jours"],
    ["", ""],
    ["[BLOCKBUSTER]", "Rotation > 0.5 -> Plus de 50% du stock vendu/semaine"],
    ["Action", "Maintenir stock eleve, produit star a forte demande"],
    ["", ""],
    ["[STOCK MORT]", "0 ventes + Stock > 10 unites -> Immobilisation"],
    ["Action", "Lancer promotion, destockage ou retour fournisseur"],
    ["", ""],
    ["[STABLE]", "Autres situations -> Gestion normale"],
    ["Action", "Surveillance continue, pas d'urgence"],
]


# RECOMMANDATIONS
recommandations = generer_recommandations(dashboard, week_num)

# EXPORT
os.makedirs(DOSSIER_SORTIE, exist_ok=True)
fichier = os.path.join(DOSSIER_SORTIE, f"Suivi_EP_W{week_num:02d}_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx")

print(f"\nGeneration du fichier Excel...")

with pd.ExcelWriter(fichier, engine='openpyxl') as writer:
    current_row = 0
    
    # KPI
    df_kpi = pd.DataFrame(list(kpi_data.items()), columns=["INDICATEURS CLES", "Valeur"])
    df_kpi.to_excel(writer, sheet_name="DASHBOARD", index=False, startrow=current_row)
    current_row += len(df_kpi) + 3
    
    # TOP
    if not top_ca.empty:
        top_ca.to_excel(writer, sheet_name="DASHBOARD", startrow=current_row, index=False)
        current_row += len(top_ca) + 3
    
    # RECO
    df_reco = pd.DataFrame(recommandations, columns=["RECOMMANDATIONS & ACTIONS", "Details"])
    df_reco.to_excel(writer, sheet_name="DASHBOARD", startrow=current_row, index=False)
    current_row += len(df_reco) + 3
    
    # GUIDE
    df_guide = pd.DataFrame(guide_lecture, columns=["GUIDE DE LECTURE", "Explication"])
    df_guide.to_excel(writer, sheet_name="DASHBOARD", startrow=current_row, index=False)
    
    # SUIVI
    cols_final = ["MARQUE", "EAN", "LIBELLE", "P.VENTE", "STOCK_EP", "BURINTEL_DEPOT", 
                  "VENTES_HEBDO", "CA_HEBDO", "ROTATION", "COUVERTURE", "STATUS"]
    disp_cols = [c for c in cols_final if c in dashboard.columns]
    
    dashboard_export = dashboard[disp_cols].sort_values("CA_HEBDO", ascending=False).copy()
    dashboard_export.to_excel(writer, sheet_name="SUIVI", index=False)
    
    # TOP CA
    if not top_ca.empty:
        top_ca.to_excel(writer, sheet_name="TOP CA", index=False)
    
    wb = writer.book
    
    # === AJOUTER LES EMOJIS MANUELLEMENT ===
    ws_dash = wb["DASHBOARD"]
    ws_suivi = wb["SUIVI"]
    
    # Emojis dans KPI
    kpi_emojis = {
        f"CA W{week_num}": f"💰 CA W{week_num}",
        "Stock EP": "📦 Stock EP",
        "Burintel Depot": "🏭 Burintel Dépôt",
        "Ventes Hebdo": "📈 Ventes Hebdo",
        "Articles Urgents": "🔴 Articles Urgents",
        "Blockbusters": "🟢 Blockbusters"
    }
    
    # Remplacer dans KPI
    for row in range(2, len(df_kpi) + 2):
        cell_value = str(ws_dash.cell(row=row, column=1).value)
        for old, new in kpi_emojis.items():
            if old in cell_value:
                ws_dash.cell(row=row, column=1).value = new
                break
    
    # Emojis dans STATUS (colonne SUIVI)
    status_emojis = {
        "URGENT": "🔴 URGENT",
        "A COMMANDE": "🟡 À COMMANDE",
        "BLOCKBUSTER": "🟢 BLOCKBUSTER",
        "STOCK MORT": "🟠 STOCK MORT",
        "STABLE": "⚪ STABLE"
    }
    
    status_col_idx = disp_cols.index("STATUS") + 1
    for row in range(2, ws_suivi.max_row + 1):
        cell = ws_suivi.cell(row=row, column=status_col_idx)
        cell_value = str(cell.value)
        if cell_value in status_emojis:
            cell.value = status_emojis[cell_value]
    
    # Ajouter emojis aux headers
    ws_dash.cell(row=1, column=1).value = "📊 INDICATEURS CLÉS"
    
    # Emojis dans TOP CA
    top_row = len(df_kpi) + 3
    if not top_ca.empty:
        ws_dash.cell(row=top_row, column=1).value = "🏆"
        # Ajouter médailles
        for i, (idx, row_data) in enumerate(top_ca.iterrows(), start=1):
            row_num = top_row + i
            rang_cell = ws_dash.cell(row=row_num, column=1)
            if i == 1:
                rang_cell.value = "🥇"
            elif i == 2:
                rang_cell.value = "🥈"
            elif i == 3:
                rang_cell.value = "🥉"
    
    # Emojis dans RECOMMANDATIONS
    reco_row = top_row + len(top_ca) + 3 if not top_ca.empty else len(df_kpi) + 3
    ws_dash.cell(row=reco_row, column=1).value = "🎯 RECOMMANDATIONS & ACTIONS"
    
    reco_emojis = {
        "VUE D'ENSEMBLE": "📊 VUE D'ENSEMBLE",
        "ACTIONS PRIORITAIRES": "🎯 ACTIONS PRIORITAIRES",
        "OPPORTUNITES": "💰 OPPORTUNITÉS",
        "PROBLEMES A RESOUDRE": "⚠️ PROBLÈMES À RÉSOUDRE",
        "PLAN D'ACTION CETTE SEMAINE": "📋 PLAN D'ACTION CETTE SEMAINE",
        "[URGENT]": "🔴 URGENT",
        "[A COMMANDE]": "🟡 À COMMANDER",
        "[BLOCKBUSTERS]": "🟢 BLOCKBUSTERS",
        "[STOCK MORT]": "🟠 STOCK MORT"
    }
    
    for row in range(reco_row + 1, reco_row + len(df_reco) + 1):
        cell_value = str(ws_dash.cell(row=row, column=1).value)
        for old, new in reco_emojis.items():
            if old in cell_value:
                ws_dash.cell(row=row, column=1).value = cell_value.replace(old, new)
                break
    
    # Emojis dans GUIDE
    guide_row = reco_row + len(df_reco) + 3
    ws_dash.cell(row=guide_row, column=1).value = "📚 GUIDE DE LECTURE"
    
    guide_emojis = {
        "FORMULES UTILISEES": "📐 FORMULES UTILISÉES",
        "SIGNIFICATION DES STATUS": "🎯 SIGNIFICATION DES STATUS",
        "[URGENT]": "🔴 URGENT",
        "[A COMMANDE]": "🟡 À COMMANDE",
        "[BLOCKBUSTER]": "🟢 BLOCKBUSTER",
        "[STOCK MORT]": "🟠 STOCK MORT",
        "[STABLE]": "⚪ STABLE"
    }
    
    for row in range(guide_row + 1, guide_row + len(df_guide) + 1):
        cell_value = str(ws_dash.cell(row=row, column=1).value)
        for old, new in guide_emojis.items():
            if old in cell_value:
                ws_dash.cell(row=row, column=1).value = cell_value.replace(old, new)
                break
    
    # === FORMATAGE ===
    appliquer_style_header(ws_dash, 1, "1F4E78")
    for row in range(2, len(df_kpi) + 2):
        for col in range(1, 3):
            cell = ws_dash.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="E7F0F8" if row % 2 == 0 else "D9E8F5", 
                                   end_color="E7F0F8" if row % 2 == 0 else "D9E8F5", 
                                   fill_type="solid")
            cell.font = Font(bold=True, size=11)
    
    if not top_ca.empty:
        appliquer_style_header(ws_dash, top_row, "FF6B35")
        for row in range(top_row + 1, top_row + len(top_ca) + 1):
            for col in range(1, 7):
                cell = ws_dash.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="FFF2CC" if row % 2 == 0 else "FFE6CC",
                                       end_color="FFF2CC" if row % 2 == 0 else "FFE6CC",
                                       fill_type="solid")
    
    appliquer_style_header(ws_dash, reco_row, "00B050")
    for row in range(reco_row + 1, reco_row + len(df_reco) + 1):
        cell_a = ws_dash.cell(row=row, column=1)
        cell_b = ws_dash.cell(row=row, column=2)
        
        if "🔴" in str(cell_a.value):
            cell_a.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            cell_a.font = Font(bold=True, size=11, color="C00000")
        elif "🟡" in str(cell_a.value):
            cell_a.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
            cell_a.font = Font(bold=True, size=11, color="E67E22")
        elif "🟢" in str(cell_a.value):
            cell_a.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            cell_a.font = Font(bold=True, size=11, color="27AE60")
        elif "🟠" in str(cell_a.value):
            cell_a.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            cell_a.font = Font(bold=True, size=11, color="D35400")
        elif any(emoji in str(cell_a.value) for emoji in ["📊", "🎯", "💰", "⚠️", "📋"]):
            cell_a.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            cell_a.font = Font(bold=True, size=11, color="000000")
        else:
            cell_a.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        cell_b.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell_b.alignment = Alignment(wrap_text=True, vertical="top")
    
    appliquer_style_header(ws_dash, guide_row, "4472C4")
    for row in range(guide_row + 1, guide_row + len(df_guide) + 1):
        cell_a = ws_dash.cell(row=row, column=1)
        cell_b = ws_dash.cell(row=row, column=2)
        
        if any(keyword in str(cell_a.value) for keyword in ["ROTATION", "COUVERTURE", "CA HEBDO", "🔴", "🟡", "🟢", "🟠", "⚪", "📐", "🎯"]):
            cell_a.fill = PatternFill(start_color="E7E6F7", end_color="E7E6F7", fill_type="solid")
            cell_a.font = Font(bold=True, size=10, color="4472C4")
        else:
            cell_a.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        cell_b.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell_b.alignment = Alignment(wrap_text=True)
    
    ws_dash.column_dimensions['A'].width = 35
    ws_dash.column_dimensions['B'].width = 70
    
    # SUIVI
    appliquer_style_header(ws_suivi, 1, "00B050")
    
    for row in range(2, ws_suivi.max_row + 1):
        for col in range(1, len(disp_cols) + 1):
            cell = ws_suivi.cell(row=row, column=col)
            
            cell.fill = PatternFill(start_color="E2F0D9" if row % 2 == 0 else "F2FFED",
                                   end_color="E2F0D9" if row % 2 == 0 else "F2FFED",
                                   fill_type="solid")
            
            if disp_cols[col-1] in ["STOCK_EP", "BURINTEL_DEPOT", "VENTES_HEBDO", "ROTATION", "COUVERTURE"]:
                cell.alignment = Alignment(horizontal="center")
    
    status_col_letter = get_column_letter(status_col_idx)
    appliquer_couleur_status(ws_suivi, status_col_letter, 2, ws_suivi.max_row)
    
    ws_suivi.column_dimensions['A'].width = 15
    ws_suivi.column_dimensions['B'].width = 15
    ws_suivi.column_dimensions['C'].width = 40
    ws_suivi.column_dimensions['D'].width = 10
    ws_suivi.column_dimensions['E'].width = 12
    ws_suivi.column_dimensions['F'].width = 14
    ws_suivi.column_dimensions['G'].width = 12
    ws_suivi.column_dimensions['H'].width = 12
    ws_suivi.column_dimensions['I'].width = 10
    ws_suivi.column_dimensions['J'].width = 12
    ws_suivi.column_dimensions['K'].width = 16
    
    # TOP CA
    if not top_ca.empty:
        ws_top = wb["TOP CA"]
        appliquer_style_header(ws_top, 1, "FFD966")
        
        for row in range(2, ws_top.max_row + 1):
            for col in range(1, ws_top.max_column + 1):
                cell = ws_top.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="FFF2CC" if row % 2 == 0 else "FFFACD",
                                       end_color="FFF2CC" if row % 2 == 0 else "FFFACD",
                                       fill_type="solid")
        
        ws_top.column_dimensions['A'].width = 5
        ws_top.column_dimensions['B'].width = 15
        ws_top.column_dimensions['C'].width = 40
        ws_top.column_dimensions['D'].width = 12
        ws_top.column_dimensions['E'].width = 10
        ws_top.column_dimensions['F'].width = 16
    
    # Renommer les onglets avec emojis
    wb["DASHBOARD"].title = "🏠 DASHBOARD"
    wb["SUIVI"].title = "📊 SUIVI"
    if "TOP CA" in wb.sheetnames:
        wb["TOP CA"].title = "🥇 TOP CA"

wb.save(fichier)

print(f"\n{'='*80}")
print(f"DASHBOARD TERMINE")
print(f"{'='*80}")
print(f"Fichier: {os.path.basename(fichier)}")
print(f"\nRESUME SEMAINE W{week_num}")
print(f"{'-'*80}")
for k, v in list(kpi_data.items()):
    print(f"   {k}: {v}")
print(f"{'='*80}")
print(f"\nDashboard avec Emojis compatibles Excel genere!")
input("\nAppuyez sur Entree pour fermer...")
